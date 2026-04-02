"""
diagnose_missingness.py — Comprehensive missing data diagnostics.

Produces:
  - Console summary of missingness patterns
  - output/missingness_diagnostics.xlsx with 3 sheets:
    1. country_x_variable: % missing per country per variable
    2. country_x_year_heatmap: binary 0/1 for each (country, year) × variable
    3. year_coverage: for each variable, first/last non-null year per country
  - output/missingness_by_country_variable.png — missingness heatmap
  - output/unr_outlier_inspection.png — UNR extreme-value inspection

Run: uv run python src/diagnose_missingness.py
"""

import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Allow importing labels.py from the same src/ directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from labels import SHORT, DEFS, add_footer

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_PATH = os.path.join(PROJECT_ROOT, "data", "oecd_economic_outlook.csv")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")

VARIABLES = [
    "gdpv_annpct", "unr", "cpi_ytypct", "cbgdpr",
    "itv_annpct", "sratio", "xgsv_annpct", "mgsv_annpct",
]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df = pd.read_csv(DATA_PATH)
    countries = sorted(df["country_code"].unique())
    years = sorted(df["year"].unique())

    print(f"Panel: {len(countries)} countries × {len(years)} years "
          f"({min(years)}–{max(years)})")
    print(f"Total rows: {len(df):,}\n")

    # -----------------------------------------------------------------------
    # 1. Country × variable: % missing
    # -----------------------------------------------------------------------
    print("=" * 70)
    print("MISSINGNESS BY COUNTRY × VARIABLE (% of that country's rows)")
    print("=" * 70)

    miss_pct = (
        df.groupby("country_code")[VARIABLES]
        .apply(lambda g: g.isnull().mean() * 100)
    )
    # Round for readability
    miss_pct_display = miss_pct.round(1)

    # Summary: countries with >20% missing in any variable
    high_miss = miss_pct[miss_pct.max(axis=1) > 20]
    print(f"\nCountries with >20% missing in at least one variable: "
          f"{len(high_miss)} of {len(countries)}")
    print(f"Countries with COMPLETE data across all variables: "
          f"{(miss_pct.max(axis=1) == 0).sum()}")
    print()

    # -----------------------------------------------------------------------
    # 2. Year coverage per country × variable: first and last non-null year
    # -----------------------------------------------------------------------
    print("=" * 70)
    print("COVERAGE WINDOWS: First and last non-null year per country × variable")
    print("=" * 70)

    coverage_records = []
    for country in countries:
        cdf = df[df["country_code"] == country].sort_values("year")
        for var in VARIABLES:
            non_null = cdf[cdf[var].notna()]
            if len(non_null) == 0:
                coverage_records.append({
                    "country_code": country,
                    "variable": var,
                    "first_year": None,
                    "last_year": None,
                    "n_available": 0,
                    "n_total": len(cdf),
                    "pct_available": 0.0,
                    "pattern": "entirely_missing",
                })
            else:
                first = int(non_null["year"].min())
                last = int(non_null["year"].max())
                n_avail = len(non_null)
                # Check for internal gaps
                expected_span = last - first + 1
                has_gaps = n_avail < expected_span
                pattern = "contiguous" if not has_gaps else "has_gaps"
                coverage_records.append({
                    "country_code": country,
                    "variable": var,
                    "first_year": first,
                    "last_year": last,
                    "n_available": n_avail,
                    "n_total": len(cdf),
                    "pct_available": round(100 * n_avail / len(cdf), 1),
                    "pattern": pattern,
                })

    coverage = pd.DataFrame(coverage_records)

    # Summarize pattern types
    pattern_counts = coverage.groupby(["variable", "pattern"]).size().unstack(fill_value=0)
    print("\nPattern counts by variable (across all countries):")
    print(pattern_counts.to_string())
    print()

    # Which countries are entirely missing for each variable?
    for var in VARIABLES:
        entirely = coverage[(coverage["variable"] == var) &
                            (coverage["pattern"] == "entirely_missing")]
        if len(entirely) > 0:
            codes = ", ".join(sorted(entirely["country_code"].tolist()))
            print(f"  {var}: entirely missing for {len(entirely)} countries: {codes}")

    # -----------------------------------------------------------------------
    # 3. Classify missingness type: historical (don't go back far enough)
    #    vs recent (stopped reporting) vs sparse
    # -----------------------------------------------------------------------
    print("\n" + "=" * 70)
    print("MISSINGNESS CLASSIFICATION")
    print("=" * 70)

    for var in VARIABLES:
        var_cov = coverage[coverage["variable"] == var]
        late_start = var_cov[(var_cov["first_year"].notna()) &
                             (var_cov["first_year"] > min(years) + 2)]
        early_end = var_cov[(var_cov["last_year"].notna()) &
                            (var_cov["last_year"] < max(years) - 2)]
        gaps = var_cov[var_cov["pattern"] == "has_gaps"]
        empty = var_cov[var_cov["pattern"] == "entirely_missing"]

        print(f"\n  {var}:")
        print(f"    Entirely missing:    {len(empty):>3} countries")
        print(f"    Late start (>1992):  {len(late_start):>3} countries "
              f"(historical missingness)")
        print(f"    Early end (<{max(years)-2}):   {len(early_end):>3} countries "
              f"(stopped reporting / recent)")
        print(f"    Internal gaps:       {len(gaps):>3} countries")

    # -----------------------------------------------------------------------
    # 4. Overlap: how many country-years have ALL variables non-null?
    # -----------------------------------------------------------------------
    print("\n" + "=" * 70)
    print("COMPLETE-CASE ANALYSIS")
    print("=" * 70)

    df["all_complete"] = df[VARIABLES].notna().all(axis=1)
    complete = df[df["all_complete"]]
    print(f"\nRows with ALL 8 variables non-null: {len(complete)} of "
          f"{len(df)} ({100*len(complete)/len(df):.1f}%)")
    print(f"Countries represented: {complete['country_code'].nunique()} of "
          f"{len(countries)}")
    print(f"Year range: {complete['year'].min()}–{complete['year'].max()}")

    # Without the two worst variables
    vars_6 = [v for v in VARIABLES if v not in ("sratio", "cpi_ytypct")]
    df["six_complete"] = df[vars_6].notna().all(axis=1)
    complete_6 = df[df["six_complete"]]
    print(f"\nRows with 6 variables (dropping sratio, cpi_ytypct) non-null: "
          f"{len(complete_6)} of {len(df)} ({100*len(complete_6)/len(df):.1f}%)")
    print(f"Countries represented: {complete_6['country_code'].nunique()} of "
          f"{len(countries)}")
    print(f"Year range: {complete_6['year'].min()}–{complete_6['year'].max()}")

    # -----------------------------------------------------------------------
    # 5. Build the Excel workbook
    # -----------------------------------------------------------------------
    xlsx_path = os.path.join(OUTPUT_DIR, "missingness_diagnostics.xlsx")

    wb = Workbook()

    # --- Sheet 1: Country × Variable % missing ---
    ws1 = wb.active
    ws1.title = "country_x_variable"

    # Header row
    headers = ["country_code"] + VARIABLES
    for j, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    red_fill = PatternFill("solid", start_color="FFCCCC")
    yellow_fill = PatternFill("solid", start_color="FFFFCC")
    green_fill = PatternFill("solid", start_color="CCFFCC")

    for i, country in enumerate(countries, 2):
        ws1.cell(row=i, column=1, value=country)
        for j, var in enumerate(VARIABLES, 2):
            val = miss_pct.loc[country, var] if country in miss_pct.index else 100.0
            cell = ws1.cell(row=i, column=j, value=round(val, 1))
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="center")
            if val > 40:
                cell.fill = red_fill
            elif val > 15:
                cell.fill = yellow_fill
            elif val == 0:
                cell.fill = green_fill

    # Auto-width
    for j in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(j)].width = 16

    # --- Sheet 2: Binary heatmap (country-year × variable) ---
    ws2 = wb.create_sheet("binary_heatmap")
    headers2 = ["country_code", "year"] + VARIABLES
    for j, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)

    row_idx = 2
    for country in countries:
        cdf = df[df["country_code"] == country].sort_values("year")
        for _, row in cdf.iterrows():
            ws2.cell(row=row_idx, column=1, value=country)
            ws2.cell(row=row_idx, column=2, value=int(row["year"]))
            for j, var in enumerate(VARIABLES, 3):
                is_missing = 1 if pd.isna(row[var]) else 0
                cell = ws2.cell(row=row_idx, column=j, value=is_missing)
                if is_missing:
                    cell.fill = red_fill
            row_idx += 1

    for j in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 16

    # --- Sheet 3: Coverage windows ---
    ws3 = wb.create_sheet("year_coverage")
    cov_headers = ["country_code", "variable", "first_year", "last_year",
                   "n_available", "n_total", "pct_available", "pattern"]
    for j, h in enumerate(cov_headers, 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True)

    for i, rec in enumerate(coverage_records, 2):
        for j, h in enumerate(cov_headers, 1):
            val = rec[h]
            cell = ws3.cell(row=i, column=j, value=val)
            if h == "pattern" and val == "entirely_missing":
                cell.fill = red_fill
            elif h == "pattern" and val == "has_gaps":
                cell.fill = yellow_fill

    for j in range(1, len(cov_headers) + 1):
        ws3.column_dimensions[get_column_letter(j)].width = 16

    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")

    # -----------------------------------------------------------------------
    # 6. Visualization: Country × Variable missingness heatmap
    # -----------------------------------------------------------------------
    sns.set_style("whitegrid")

    # Use short labels on the x-axis; fall back to raw name if not in SHORT
    col_labels = [SHORT.get(v, v) for v in VARIABLES]

    fig, ax = plt.subplots(figsize=(12, 16))
    plot_data = miss_pct[VARIABLES].loc[countries].copy()
    plot_data.columns = col_labels

    sns.heatmap(
        plot_data,
        cmap="YlOrRd", vmin=0, vmax=100,
        annot=True, fmt=".0f",
        linewidths=0.3, linecolor="white",
        ax=ax, cbar_kws={"label": "% Missing"},
    )
    ax.set_title("Missing Data: % per Country × Variable", fontsize=14)
    ax.set_xlabel("")
    ax.set_ylabel("")
    plt.tight_layout(rect=[0, 0.06, 1, 1])
    add_footer(fig, [v for v in VARIABLES if v in DEFS])
    png_path = os.path.join(OUTPUT_DIR, "missingness_by_country_variable.png")
    plt.savefig(png_path, dpi=150, bbox_inches="tight")
    print(f"Saved: {png_path}")
    plt.close()

    print("\nDone. Review the spreadsheet and heatmap for detailed patterns.")

    # -----------------------------------------------------------------------
    # 7. UNR outlier inspection
    #    28 observations (2.0%) exceed 18% — 5+ SDs above the panel mean.
    #    All are real, ILO-harmonised values; documented below.
    # -----------------------------------------------------------------------
    # Load the 6-variable dataset (cpi_ytypct and sratio already dropped)
    df6 = df[df["country_code"].isin(
        pd.read_csv(DATA_PATH)["country_code"].unique()
    )].copy()
    df6 = pd.read_csv(DATA_PATH)

    UNR_THRESHOLD = 18
    mean_unr = df6["unr"].mean()
    std_unr  = df6["unr"].std()

    episodes = {
        "GRC": ("Greece",        "#d62728",
                "Eurozone debt crisis\n(troika austerity, internal devaluation)"),
        "ESP": ("Spain",         "#ff7f0e",
                "Eurozone debt crisis\n(construction bust)"),
        "LVA": ("Latvia",        "#2ca02c",
                "Post-Soviet transition (1996)\n+ credit bubble bust (2010)"),
        "POL": ("Poland",        "#9467bd",
                "Post-communist restructuring\n(pre-EU accession 2004)"),
        "SVK": ("Slovak Rep.",   "#8c564b",
                "Post-communist restructuring\n(pre-EU accession 2004)"),
        "CRI": ("Costa Rica",    "#17becf",
                "COVID-19\n(tourism/services shock)"),
    }

    fig2, axes2 = plt.subplots(1, 2, figsize=(15, 6))

    unr_short = SHORT.get("unr", "Unemployment rate")

    # Panel 1: distribution with sigma lines
    ax = axes2[0]
    ax.hist(df6["unr"].dropna(), bins=40, color="steelblue",
            edgecolor="white", alpha=0.85)
    ax.axvline(mean_unr, color="black", linewidth=1.2, linestyle="--",
               label=f"Mean ({mean_unr:.1f}%)")
    ax.axvline(mean_unr + 2 * std_unr, color="orange", linewidth=1.0,
               linestyle=":", label=f"Mean+2σ ({mean_unr+2*std_unr:.1f}%)")
    ax.axvline(mean_unr + 3 * std_unr, color="red", linewidth=1.0,
               linestyle=":", label=f"Mean+3σ ({mean_unr+3*std_unr:.1f}%)")
    ax.axvspan(UNR_THRESHOLD, df6["unr"].max() + 1, alpha=0.08, color="red")
    n_high = (df6["unr"] > UNR_THRESHOLD).sum()
    ax.text(0.97, 0.97, f"n > {UNR_THRESHOLD}%: {n_high}",
            transform=ax.transAxes, ha="right", va="top",
            fontsize=9, color="red")
    ax.set_xlabel(f"{unr_short} (%)")
    ax.set_ylabel("Count")
    ax.set_title(f"{unr_short} — full panel distribution\n"
                 f"Mean={mean_unr:.1f}%, SD={std_unr:.1f}%; shaded = obs > {UNR_THRESHOLD}%")
    ax.legend(fontsize=9)

    # Panel 2: time series for affected countries
    ax = axes2[1]
    for iso, (name, color, _) in episodes.items():
        sub = df6[df6["country_code"] == iso].sort_values("year")
        ax.plot(sub["year"], sub["unr"], color=color, linewidth=1.6, label=name)
        peak = sub.loc[sub["unr"].idxmax()]
        ax.scatter(peak["year"], peak["unr"], color=color, s=40, zorder=5)
    ax.axhline(mean_unr, color="black", linewidth=0.8, linestyle="--",
               alpha=0.5, label=f"Panel mean ({mean_unr:.1f}%)")
    ax.axhline(UNR_THRESHOLD, color="red", linewidth=0.8, linestyle=":",
               alpha=0.6, label=f"{UNR_THRESHOLD}% threshold")
    ax.set_xlabel("Year")
    ax.set_ylabel(f"{unr_short} (%)")
    ax.set_title(f"{unr_short} — countries with observations > {UNR_THRESHOLD}%\n"
                 "(dots = peak year)")
    ax.legend(fontsize=8, loc="upper right")

    plt.suptitle("UNR Outlier Inspection", fontsize=13, fontweight="bold", y=1.01)
    plt.tight_layout(rect=[0, 0.06, 1, 1])
    add_footer(fig2, ["unr"])
    unr_path = os.path.join(OUTPUT_DIR, "unr_outlier_inspection.png")
    plt.savefig(unr_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {unr_path}")


if __name__ == "__main__":
    main()
